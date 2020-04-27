#!/usr/bin/env python3

import openpyxl
from copy import copy
from pathlib import Path
import os
import time
import shutil

def master_eat_children(mastername, children, m, backupdir, actualdir):
    if actualdir.exists() == True:
        with open(actualdir, 'rb') as f:
            with open(backupdir, 'wb+') as f2:
                f2.write(f.read())
        os.remove(actualdir)
        masterwb = openpyxl.Workbook()
        masterwb.save(actualdir)
    else:
        masterwb = openpyxl.Workbook()
        masterwb.save(actualdir)
    masterwb = openpyxl.load_workbook(actualdir)
    for i in children:
        childwb = openpyxl.load_workbook(i) 
        eachsheet = len(childwb.sheetnames)
        print(eachsheet)
        for i in range(0, eachsheet):            
            print(m) 
            childws = childwb.worksheets[i]
            if m != 0:
                masterwb.create_sheet(title=childws.title)
            masterws = masterwb.worksheets[m]
            masterws.title = childws.title
            mr = childws.max_row 
            mc = childws.max_column
            try:
                biggestcolumn = childws.column_dimensions[chr(65)].width
            except:
                pass
            print(biggestcolumn)
            try:
                biggestrow = childws.row_dimensions[1].height
            except:
                pass
            print(biggestrow)
            for i in range (1, mr + 1): 
                for j in range (1, mc + 1): 
                    c = childws.cell(row = i, column = j) 
                    mcell = masterws.cell(row = i, column = j, value=c.value)
                    mcell.font = copy(c.font)
                    mcell.border = copy(c.border)
                    mcell.fill = copy(c.fill)
                    mcell.number_format = copy(c.number_format)
                    mcell.protection = copy(c.protection)
                    mcell.alignment = copy(c.alignment)
            auto_format_cell_width3(masterws)
            m = m + 1
    masterwb.save(actualdir)
            

def auto_format_cell_width3(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
    return

def main():
    mastername = input('Please enter the Full Path Master File: \n')
    childrenfol = input('Move excels you want to merge in single folder, enter folder here: \n')
    children = input('Please enter the names of the files you want to merge, Seperate them by , and no spaces: \n')
    children = children.split(',')
    newchildren = []
    for i in children:
        newchildren.append(childrenfol+i)
    backupdir = input('Enter the full path of the folder to backup your master copy: \n')
    m = 0
    actualdir = Path(mastername)
    backupdir = Path(mastername.strip('.xlsx') +' ' +time.ctime().replace(":","_") + '.xlsx')
    master_eat_children(mastername, newchildren, m, backupdir, actualdir)


if __name__ == "__main__":
    main()


